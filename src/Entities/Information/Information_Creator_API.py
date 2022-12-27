import re
from operator import itemgetter
from pathlib import Path
from pprint import pprint
import matplotlib.pyplot as plt
from gensim.corpora import Dictionary
from gensim.models import LdaMulticore, CoherenceModel
from openpyxl import Workbook, drawing
from openpyxl.drawing import image
from openpyxl.styles import Font
import spacy
from src.Entities.Storage.MongoDB_Connector import MongoDB_Connector


class Information_Creator_API:

    @staticmethod
    def create_excel_file(mode):
        storage = MongoDB_Connector()

        psychotherapy_document_list = storage.get_filtered_documents(mode, filter_type="psychotherapy")
        psychotherapy_file_name = mode + "_psychotherapy"
        Information_Creator_API.create_information(psychotherapy_document_list, psychotherapy_file_name)

        psychology_document_list = storage.get_filtered_documents(mode, filter_type="psychology_subject")
        psychology_file_name = mode + "_psychology"
        Information_Creator_API.create_information(psychology_document_list, psychology_file_name)

        overall_document_list = storage.get_all_documents(mode)
        overall_file_name = mode
        Information_Creator_API.create_information(overall_document_list, overall_file_name)


    @staticmethod
    def create_information(document_list, file_name):
        print(f"Creating information about {file_name}")

        print(f"Loading the dataset")

        nlp = spacy.load("en_core_web_sm")

        totalSearchResults = 0

        searchResultsWithTags = 0
        searchResultsWithoutTags = 0
        tag_list = []

        searchResultsWithSubjects = 0
        searchResultsWithoutSubjects = 0
        subject_list = []

        searchResultsWithTitle = 0
        searchResultsWithoutTitles = 0
        title_list = []

        searchResultsWithDescription = 0
        searchResultsWithoutDescription = 0
        description_list = []


        for document_index, document_object in enumerate(document_list):

            if document_index % 100 == 0:
                percentage = format((document_index + 1) / len(document_list), '.4f')
                print(f"Dealing with document {document_index}/{len(document_list)} - {percentage}%")

            # if document_index == 1000:
            #     break

            totalSearchResults += 1

            document_tag_list = document_object["tag_list"]

            if document_tag_list:
                searchResultsWithTags += 1

                temp_tag_list = []

                for tag in document_tag_list:
                    split_tag_list = re.split(r"[;,#]", tag)
                    for split_tag in split_tag_list:
                        lower_tag = split_tag.lower()
                        temp_tag_list.append(lower_tag)
                tag_list.append(temp_tag_list)

            else:
                searchResultsWithoutTags += 1

            document_subject_list = document_object["subject_list"]

            if document_subject_list:
                searchResultsWithSubjects += 1

                temp_subject_list = []
                for subject in document_subject_list:
                    lower_subject = subject.lower()
                    temp_subject_list.append(lower_subject)
                subject_list.append(temp_subject_list)

            else:
                searchResultsWithoutSubjects += 1


            document_title_list = document_object["title"].split(" ")
            lower_title_list = [x.lower() for x in document_title_list]
            lower_title_str = " ".join(lower_title_list)
            nlp_title_list = nlp(lower_title_str)
            title_lemma_list = [token.lemma_ for token in nlp_title_list if (not token.is_stop) and (token.is_alpha)]

            if title_lemma_list:
                searchResultsWithTitle += 1
                title_list.append(title_lemma_list)
            else:
                searchResultsWithoutTitles += 1


            document_description_list = document_object["description"].split(" ")
            lower_description_list = [x.lower() for x in document_description_list]
            lower_description_str = " ".join(lower_description_list)
            nlp_description_list = nlp(lower_description_str)
            description_lemma_list = [token.lemma_ for token in nlp_description_list if (not token.is_stop) and (token.is_alpha)]


            if description_lemma_list:
                searchResultsWithDescription += 1
                description_list.append(description_lemma_list)

            else:
                searchResultsWithoutDescription += 1

        print(f"Determining topics")

        topic_list = [
            ("tag", tag_list, searchResultsWithTags, searchResultsWithoutTags, totalSearchResults),
            ("subject", subject_list, searchResultsWithSubjects, searchResultsWithoutSubjects, totalSearchResults),
            ("title", title_list, searchResultsWithTitle, searchResultsWithoutTitles, totalSearchResults),
            ("description", description_list, searchResultsWithDescription, searchResultsWithoutDescription, totalSearchResults)
        ]

        wb = Workbook()
        dest_filename = f"overall_information_{file_name}.xlsx"
        root_path = Path(__file__).parents[3]
        out_folder_path = root_path.joinpath("resources", "out")
        dest_path = out_folder_path.joinpath(dest_filename)

        for topic_name, information_list, documents_with_token, documents_without_token, total_document_count in topic_list:
            Information_Creator_API.determine_topics(topic_name, file_name, information_list, wb , documents_with_token,
                                                     documents_without_token, total_document_count)

        if 'Sheet' in wb.sheetnames:
            wb.remove(wb['Sheet'])

        wb.save(filename=dest_path)

        print("Finished determining topics")


    @staticmethod
    def determine_topics(topic_name, file_name, information_list, work_book, documents_with_token, documents_without_token,
                         total_document_count):

        root_path = Path(__file__).parents[3]
        out_folder_path = root_path.joinpath("resources", "out")
        images_folder_path = out_folder_path.joinpath("images")

        print(f"Creating {topic_name} corpus")
        information_dic = Dictionary(information_list)
        information_dic.filter_extremes(no_below=5, no_above=0.5, keep_n=1000)
        information_corpus = [information_dic.doc2bow(doc) for doc in information_list]

        information_count_list = []
        for token_str, token_id in information_dic.token2id.items():
            count = information_dic.cfs[token_id]
            information_count_list.append((token_str, count))
        information_count_list.sort(key=lambda x: x[1], reverse=True)

        image_name = f"{file_name}_{topic_name}_image.jpg"
        information_image_path = images_folder_path.joinpath(image_name)
        information_min_topics = 1
        information_max_topics = 20
        information_topics = []
        information_score = []
        for i in range(information_min_topics, information_max_topics + 1, 1):
            print(f"Creating {topic_name} topic model with {i} topics out of {information_max_topics}")

            information_lda_model = LdaMulticore(corpus=information_corpus, id2word=information_dic, iterations=10, num_topics=i,
                                           workers=6,
                                           passes=10, random_state=100)
            information_cm = CoherenceModel(model=information_lda_model, corpus=information_corpus, dictionary=information_dic,
                                      coherence='u_mass')
            information_topics.append(i)
            information_score.append(information_cm.get_coherence())
        _ = plt.plot(information_topics, information_score)
        _ = plt.xlabel("# topics")
        _ = plt.ylabel("u_mass score")
        _ = plt.title(f"Number of topics to coherence relationship for \"{topic_name.capitalize()}\" attribute")

        zipped_result_topics = list(zip(information_topics, information_score))
        max_topics_information, max_score_information = max(zipped_result_topics, key=itemgetter(1))

        plt.plot(max_topics_information, max_score_information, 'ro')
        plt.annotate(text="best coherence", xy=(max_topics_information, max_score_information), xytext=(max_topics_information + 0.3, max_score_information - 0.02))

        plt.xticks(ticks=[x for x in range(2, information_max_topics + 1, 2)])
        plt.savefig(information_image_path)
        print(f"{topic_name}: max coherence score {max_score_information} for {max_topics_information} topics")

        print(f"Generating final {topic_name} topic model with {max_topics_information} topics")
        final_topic_model = LdaMulticore(corpus=information_corpus, id2word=information_dic, iterations=10,
                                         num_topics=max_topics_information, workers=6, passes=10, random_state=100)
        auto_topics = final_topic_model.print_topics()
        pprint(final_topic_model.print_topics())


        print(f"Finished creating information for {topic_name}")
        print("Storing information")

        work_sheet_name = f"{topic_name}_analysis"
        work_sheet = work_book.create_sheet(work_sheet_name)

        row_count = 2
        for information, count in information_count_list:
            work_sheet.cell(row=row_count, column=1).value = information
            work_sheet.cell(row=row_count, column=2).value = count
            row_count += 1

        work_sheet.cell(row=1, column=1).value = f"{topic_name}"
        work_sheet.cell(row=1, column=1).font = Font(bold=True)
        work_sheet.cell(row=1, column=2).value = "count"
        work_sheet.cell(row=1, column=2).font = Font(bold=True)

        work_sheet.cell(row=1, column=4).value = f"documents without {topic_name}"
        work_sheet.cell(row=1, column=4).font = Font(bold=True)
        work_sheet.cell(row=1, column=5).value = documents_with_token
        work_sheet.cell(row=2, column=4).value = f"documents with {topic_name}"
        work_sheet.cell(row=2, column=4).font = Font(bold=True)
        work_sheet.cell(row=2, column=5).value = documents_without_token
        work_sheet.cell(row=3, column=4).value = "total documents"
        work_sheet.cell(row=3, column=4).font = Font(bold=True)
        work_sheet.cell(row=3, column=5).value = total_document_count

        work_sheet.cell(row=1, column=7).value = "topic name"
        work_sheet.cell(row=1, column=7).font = Font(bold=True)
        work_sheet.cell(row=1, column=8).value = "topic composition"
        work_sheet.cell(row=1, column=8).font = Font(bold=True)

        row_count = 2
        for topic_name, topic_structure in auto_topics:
            work_sheet.cell(row=row_count, column=8).value = topic_structure
            row_count += 1

        row_count += 3
        img = image.Image(str(information_image_path))
        work_sheet.add_image(img, f"G{row_count}")
