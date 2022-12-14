import collections
import winsound
from openpyxl import Workbook
from openpyxl.styles import Font
import spacy


class Information_Creator:

    @staticmethod
    def create_information(searchResultList, search_term):
        print("Creating information")

        nlp = spacy.load("en_core_web_sm")

        totalSearchResults = 0

        searchResultsWithTags = 0
        searchResultsWithoutTags = 0
        tag_count_dic = collections.defaultdict(int)

        searchResultsWithTitle = 0
        searchResultsWithoutTitles = 0
        title_list = []

        searchResultsWithDescription = 0
        searchResultsWithoutDescription = 0
        description_list = []

        for searchResult in searchResultList:
            totalSearchResults += 1

            if len(searchResult.tags) > 0:
                searchResultsWithTags += 1

                for tag in searchResult.tags:
                    tag_count_dic[tag] += 1

            else:
                searchResultsWithoutTags += 1

            if searchResult.title:
                searchResultsWithTitle += 1
                title_str = searchResult.title
                if title_str[-1] != ".":
                    title_str += "."
                title_list.append(title_str)
            else:
                searchResultsWithoutTitles += 1

            if searchResult.description:
                searchResultsWithDescription += 1
                description_str = searchResult.description
                if description_str[-1] != ".":
                    description_str += "."
                description_list.append(description_str)
            else:
                searchResultsWithoutDescription += 1

        tag_count_list = [(k, v) for k, v in tag_count_dic.items()]
        tag_count_list.sort(key=lambda x: x[1], reverse=True)

        description_str = " ".join(description_list)
        title_str = " ".join(title_list)

        description_str = description_str.lower()
        title_str = title_str.lower()

        description_tokens = nlp(description_str)
        title_tokens = nlp(title_str)

        description_lemma_list = [token.lemma_ for token in description_tokens if
                                  (not token.is_stop) and (token.is_alpha)]
        title_lemma_list = [token.lemma_ for token in title_tokens if (not token.is_stop) and (token.is_alpha)]

        description_counter = collections.Counter(description_lemma_list)
        title_counter = collections.Counter(title_lemma_list)

        description_count_list = [(k, v) for k, v in description_counter.items()]
        description_count_list.sort(key=lambda x: x[1], reverse=True)

        title_count_list = [(k, v) for k, v in title_counter.items()]
        title_count_list.sort(key=lambda x: x[1], reverse=True)

        print("Finished creating information")
        print("Storing information")

        wb = Workbook()
        dest_filename = f"{search_term}_information.xlsx"

        ws_tag = wb.active
        ws_tag.title = "Tag_Analysis"

        row_count = 2
        for tag, count in tag_count_list:
            ws_tag.cell(row=row_count, column=1).value = tag
            ws_tag.cell(row=row_count, column=2).value = count
            row_count += 1

        ws_tag.cell(row=1, column=1).value = "Tag"
        ws_tag.cell(row=1, column=1).font = Font(bold=True)
        ws_tag.cell(row=1, column=2).value = "Count"
        ws_tag.cell(row=1, column=2).font = Font(bold=True)

        ws_tag.cell(row=2, column=4).value = "Files without tags"
        ws_tag.cell(row=2, column=4).font = Font(bold=True)
        ws_tag.cell(row=2, column=5).value = searchResultsWithoutTags
        ws_tag.cell(row=3, column=4).value = "Files with tags"
        ws_tag.cell(row=3, column=4).font = Font(bold=True)
        ws_tag.cell(row=3, column=5).value = searchResultsWithTags
        ws_tag.cell(row=4, column=4).value = "Total"
        ws_tag.cell(row=4, column=4).font = Font(bold=True)
        ws_tag.cell(row=4, column=5).value = totalSearchResults

        # Title Analysis

        ws_title = wb.create_sheet("Title_Analysis")

        row_count = 2
        for title, count in title_count_list:
            ws_title.cell(row=row_count, column=1).value = title
            ws_title.cell(row=row_count, column=2).value = count
            row_count += 1

        ws_title.cell(row=1, column=1).value = "Title"
        ws_title.cell(row=1, column=1).font = Font(bold=True)
        ws_title.cell(row=1, column=2).value = "Count"
        ws_title.cell(row=1, column=2).font = Font(bold=True)

        ws_title.cell(row=2, column=4).value = "Files without titles"
        ws_title.cell(row=2, column=4).font = Font(bold=True)
        ws_title.cell(row=2, column=5).value = searchResultsWithoutTitles
        ws_title.cell(row=3, column=4).value = "Files with titles"
        ws_title.cell(row=3, column=4).font = Font(bold=True)
        ws_title.cell(row=3, column=5).value = searchResultsWithTitle
        ws_title.cell(row=4, column=4).value = "Total"
        ws_title.cell(row=4, column=4).font = Font(bold=True)
        ws_title.cell(row=4, column=5).value = totalSearchResults

        # Description Analysis

        ws_description = wb.create_sheet("Description_Analysis")

        row_count = 2
        for description, count in description_count_list:
            ws_description.cell(row=row_count, column=1).value = description
            ws_description.cell(row=row_count, column=2).value = count
            row_count += 1

        ws_description.cell(row=1, column=1).value = "Description"
        ws_description.cell(row=1, column=1).font = Font(bold=True)
        ws_description.cell(row=1, column=2).value = "Count"
        ws_description.cell(row=1, column=2).font = Font(bold=True)

        ws_description.cell(row=2, column=4).value = "Files without descriptions"
        ws_description.cell(row=2, column=4).font = Font(bold=True)
        ws_description.cell(row=2, column=5).value = searchResultsWithoutDescription
        ws_description.cell(row=3, column=4).value = "Files with descriptions"
        ws_description.cell(row=3, column=4).font = Font(bold=True)
        ws_description.cell(row=3, column=5).value = searchResultsWithDescription
        ws_description.cell(row=4, column=4).value = "Total"
        ws_description.cell(row=4, column=4).font = Font(bold=True)
        ws_description.cell(row=4, column=5).value = totalSearchResults

        wb.save(filename=dest_filename)

        print("Finished storing information")

        duration = 1000
        freq = 440
        winsound.Beep(freq, duration)
