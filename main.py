import collections

import winsound
from bs4 import BeautifulSoup, Comment
from openpyxl import Workbook
from openpyxl.styles import Font
from pyppeteer import launch
import asyncio
import spacy

from Search_Result import Search_Result


def create_information(searchResultList, search_term):
    print("Creating information")

    nlp = spacy.load("en_core_web_sm")

    totalSearchResults = 0

    searchResultsWithTags = 0
    searchResultsWithoutTags = 0
    tag_count_dic = collections.defaultdict(int)

    searchResultsWithTitle = 0
    searchResultsWithoutTitles = 0
    title_list = []

    searchResultsWithDescription = 0
    searchResultsWithoutDescription = 0
    description_list = []

    for searchResult in searchResultList:
        totalSearchResults += 1

        if len(searchResult.tags) > 0:
            searchResultsWithTags += 1

            for tag in searchResult.tags:
                tag_count_dic[tag] += 1

        else:
            searchResultsWithoutTags += 1

        if searchResult.title:
            searchResultsWithTitle += 1
            title_str = searchResult.title
            if title_str[-1] != ".":
                title_str += "."
            title_list.append(title_str)
        else:
            searchResultsWithoutTitles += 1

        if searchResult.description:
            searchResultsWithDescription += 1
            description_str = searchResult.description
            if description_str[-1] != ".":
                description_str += "."
            description_list.append(description_str)
        else:
            searchResultsWithoutDescription += 1

    tag_count_list = [(k, v) for k,v in tag_count_dic.items()]
    tag_count_list.sort(key= lambda x: x[1], reverse=True)

    description_str = " ".join(description_list)
    title_str = " ".join(title_list)

    description_str = description_str.lower()
    title_str = title_str.lower()

    description_tokens = nlp(description_str)
    title_tokens = nlp(title_str)

    description_lemma_list = [token.lemma_ for token in description_tokens if (not token.is_stop) and (token.is_alpha)]
    title_lemma_list = [token.lemma_ for token in title_tokens if (not token.is_stop) and (token.is_alpha)]

    description_counter = collections.Counter(description_lemma_list)
    title_counter = collections.Counter(title_lemma_list)

    description_count_list = [(k, v) for k,v in description_counter.items()]
    description_count_list.sort(key= lambda x: x[1], reverse=True)

    title_count_list = [(k, v) for k,v in title_counter.items()]
    title_count_list.sort(key= lambda x: x[1], reverse=True)

    print("Finished creating information")
    print("Storing information")

    wb = Workbook()
    dest_filename = f"{search_term}_information.xlsx"

    ws_tag = wb.active
    ws_tag.title = "Tag_Analysis"

    row_count = 2
    for tag, count in tag_count_list:
        ws_tag.cell(row=row_count, column=1).value = tag
        ws_tag.cell(row=row_count, column=2).value = count
        row_count += 1

    ws_tag.cell(row=1, column=1).value = "Tag"
    ws_tag.cell(row=1, column=1).font = Font(bold=True)
    ws_tag.cell(row=1, column=2).value = "Count"
    ws_tag.cell(row=1, column=2).font = Font(bold=True)

    ws_tag.cell(row=2, column=4).value = "Files without tags"
    ws_tag.cell(row=2, column=4).font = Font(bold=True)
    ws_tag.cell(row=2, column=5).value = searchResultsWithoutTags
    ws_tag.cell(row=3, column=4).value = "Files with tags"
    ws_tag.cell(row=3, column=4).font = Font(bold=True)
    ws_tag.cell(row=3, column=5).value = searchResultsWithTags
    ws_tag.cell(row=4, column=4).value = "Total"
    ws_tag.cell(row=4, column=4).font = Font(bold=True)
    ws_tag.cell(row=4, column=5).value = totalSearchResults

    # Title Analysis

    ws_title = wb.create_sheet("Title_Analysis")

    row_count = 2
    for title, count in title_count_list:
        ws_title.cell(row=row_count, column=1).value = title
        ws_title.cell(row=row_count, column=2).value = count
        row_count += 1

    ws_title.cell(row=1, column=1).value = "Title"
    ws_title.cell(row=1, column=1).font = Font(bold=True)
    ws_title.cell(row=1, column=2).value = "Count"
    ws_title.cell(row=1, column=2).font = Font(bold=True)

    ws_title.cell(row=2, column=4).value = "Files without titles"
    ws_title.cell(row=2, column=4).font = Font(bold=True)
    ws_title.cell(row=2, column=5).value = searchResultsWithoutTitles
    ws_title.cell(row=3, column=4).value = "Files with titles"
    ws_title.cell(row=3, column=4).font = Font(bold=True)
    ws_title.cell(row=3, column=5).value = searchResultsWithTitle
    ws_title.cell(row=4, column=4).value = "Total"
    ws_title.cell(row=4, column=4).font = Font(bold=True)
    ws_title.cell(row=4, column=5).value = totalSearchResults

    # Description Analysis

    ws_description = wb.create_sheet("Description_Analysis")

    row_count = 2
    for description, count in description_count_list:
        ws_description.cell(row=row_count, column=1).value = description
        ws_description.cell(row=row_count, column=2).value = count
        row_count += 1

    ws_description.cell(row=1, column=1).value = "Description"
    ws_description.cell(row=1, column=1).font = Font(bold=True)
    ws_description.cell(row=1, column=2).value = "Count"
    ws_description.cell(row=1, column=2).font = Font(bold=True)

    ws_description.cell(row=2, column=4).value = "Files without descriptions"
    ws_description.cell(row=2, column=4).font = Font(bold=True)
    ws_description.cell(row=2, column=5).value = searchResultsWithoutDescription
    ws_description.cell(row=3, column=4).value = "Files with descriptions"
    ws_description.cell(row=3, column=4).font = Font(bold=True)
    ws_description.cell(row=3, column=5).value = searchResultsWithDescription
    ws_description.cell(row=4, column=4).value = "Total"
    ws_description.cell(row=4, column=4).font = Font(bold=True)
    ws_description.cell(row=4, column=5).value = totalSearchResults

    wb.save(filename=dest_filename)

    print("Finished storing information")

    duration = 1000
    freq = 440
    winsound.Beep(freq, duration)


async def main(search_term):
    print(f"Crawling osf.io for the term: {search_term}")

    browser = await launch()
    page = await browser.newPage()

    searchResultList = []

    await page.goto(f'https://osf.io/search/?q={search_term}&page=1')
    html = await page.content()
    soup = BeautifulSoup(html, 'html.parser')
    body = soup.body
    for element in body(text=lambda text: isinstance(text, Comment)):
        element.extract()

    pagger = body.find_all("ul", class_="pager")[0]
    max_page_number = int(pagger.findChild("span").text.split(" of ")[1])

    for osf_page in range(1, max_page_number):

        print_str = f"Checking Page {osf_page} / {max_page_number}"
        print(print_str)

        search_results = body.find_all("div", class_="search-result")

        for result in search_results:

            title = None
            file_name = None
            registration_date = None
            description = None
            contributors = []
            tags = []

            h4_child = result.findChild("h4")
            if h4_child:
                a_child = h4_child.findChild("a")
                if a_child:
                    title = a_child.text
                    file_name = a_child["href"]


            span_children = result.findChildren("span")
            if span_children:
                for span in span_children:
                    data_bind = span["data-bind"]

                    if data_bind == "text: 'Date Registered: ' + dateRegistered['local'], tooltip: {title: dateRegistered['utc']}":
                        registration_date = span.text.split("Date Registered: ")[1]

                    elif data_bind == "foreach: contributors":
                        a_children = span.findChildren("a")
                        if a_children:
                            for a_child in a_children:
                                contributors.append(a_child.text)

                    elif data_bind == "fitText: {text: description, length: 500}":
                        description = span.text

                    elif data_bind == "text: $data":
                        tags.append(span.text)

            newSearchResult = Search_Result(title, file_name, registration_date, description, contributors, tags)
            searchResultList.append(newSearchResult)

        next_page_number = osf_page + 1
        if next_page_number == max_page_number + 1:
        #if next_page_number == 4:
            print(f"Finished crawling osf.io for the term: {search_term}")
            break
        else:
            await page.goto(f'https://osf.io/search/?q={search_term}&page={next_page_number}')
            html = await page.content()
            soup = BeautifulSoup(html, 'html.parser')
            body = soup.body
            for element in body(text=lambda text: isinstance(text, Comment)):
                element.extract()

    create_information(searchResultList, search_term)

asyncio.get_event_loop().run_until_complete(main("placebo"))
